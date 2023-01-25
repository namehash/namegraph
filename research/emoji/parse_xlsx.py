from argparse import ArgumentParser
from typing import Any
from collections import defaultdict
from string import ascii_lowercase, digits
from itertools import count, combinations_with_replacement, chain
from copy import deepcopy
import json

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def get_color(color_hex: str) -> str | None:
    color_map = {
        'FFB6D7A8': 'green',
        'FFD9EAD3': 'green',
        'FFEA9999': 'red',
        'FFE6B8AF': 'red',
        'FFDD7E6B': 'red',
        'FFF4CCCC': 'red',
        'FFFFE599': 'yellow',
        '00000000': 'yellow'
    }
    return color_map[color_hex]


def parser_xlsx(filepath: str):
    wb = load_workbook(filepath)
    ws = wb['Words']

    colors = set()
    name2emojis = dict()
    for row in count(2):
        name = ws[get_column_letter(2) + str(row)].value

        if not name:
            break

        emojis = defaultdict(list)
        for column in count(4):
            coords = get_column_letter(column) + str(row)

            if not ws[coords].value:
                break

            color_hex = ws[coords].fill.start_color.index
            colors.add(color_hex)

            color = get_color(color_hex)
            if color:
                emojis[color].append(ws[coords].value)

        name2emojis[name] = emojis

    return name2emojis


def pad_empty_short_words(name2emojis: dict[str, Any]) -> dict[str, Any]:
    chars = ascii_lowercase + digits

    name2emojis = deepcopy(name2emojis)
    for short_word in chain(combinations_with_replacement(chars, 1),
                            combinations_with_replacement(chars, 2),
                            combinations_with_replacement(chars, 3)):

        short_word = ''.join(short_word)

        if short_word not in name2emojis:
            name2emojis[short_word] = dict()

    return name2emojis


if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument('sheet', help='path to the xlsx file')
    parser.add_argument('output', help='path to the json output file')
    parser.add_argument('--pad_empty_short_words', action='store_true', help='sets all the other up to 3 character length words as such that generate no emojis')
    args = parser.parse_args()

    name2emojis = parser_xlsx(args.sheet)
    if args.pad_empty_short_words:
        name2emojis = pad_empty_short_words(name2emojis)

    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(name2emojis, f, indent=2, ensure_ascii=False)
