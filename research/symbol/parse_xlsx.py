from argparse import ArgumentParser
from collections import defaultdict
from itertools import count
import json

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import myunicode


def parser_xlsx(filepath: str):
    wb = load_workbook(filepath)

    symbol2names = dict()
    for worksheet_name in ['other', 'math', 'currency']:
        ws = wb[worksheet_name]

        for row in count(1):
            symbol = ws[get_column_letter(2) + str(row)].value

            if not symbol:
                break

            names = []
            for column in count(4):
                coords = get_column_letter(column) + str(row)

                if not ws[coords].value:
                    break

                if ws[coords].font.bold:
                    names.append(ws[coords].value)

            if names:
                symbol2names[symbol] = names

    return symbol2names


def filter_out_unnormalized(symbol2names: dict[str, list[str]]) -> dict[str, list[str]]:
    filtered = dict()
    for symbol, value in symbol2names.items():
        if myunicode.is_ens_normalized(symbol) and myunicode.script_of(symbol) == 'Common':
            filtered[symbol] = value
        else:
            print(f'Filtered out {repr(symbol)}, script - {myunicode.script_of(symbol)}')
    return filtered


def invert_mapping(symbol2names: dict[str, list[str]]) -> dict[str, list[str]]:
    name2symbols = defaultdict(list)
    for symbol, names in symbol2names.items():
        for name in names:
            name2symbols[name].append(symbol)

    return dict(name2symbols)


def sort_mapping(name2symbols: dict[str, list[str]], freqs: dict[str, int]) -> dict[str, list[str]]:
    name2sorted_symbols = dict()
    for name, symbols in name2symbols.items():
        name2sorted_symbols[name] = sorted(symbols, key=lambda x: freqs.get(x, 0), reverse=True)
    return name2sorted_symbols


if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument('sheet', help='path to the xlsx file')
    parser.add_argument('freqs', help='path to the json file with symbol frequencies')
    parser.add_argument('output', help='path to the json output file')
    args = parser.parse_args()

    symbol2names = parser_xlsx(args.sheet)
    filtered_symbol2names = filter_out_unnormalized(symbol2names)
    name2symbols = invert_mapping(filtered_symbol2names)

    with open(args.freqs, 'r', encoding='utf-8') as f:
        freqs = json.load(f)

    name2sorted_symbols = sort_mapping(name2symbols, freqs)

    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(name2sorted_symbols, f, indent=2, ensure_ascii=False)
